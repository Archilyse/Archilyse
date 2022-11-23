import csv
import re
from io import StringIO
from typing import Any, Mapping, Optional, Union
from zipfile import BadZipFile

from marshmallow import Schema, ValidationError, fields
from marshmallow_enum import EnumField
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from common_utils.constants import SUPPORTED_LANGUAGES, SUPPORTED_OUTPUT_FILES


class CapitalizedStr(fields.Str):
    def _deserialize(self, value, attr, data, **kwargs) -> Any:
        deserialized_str = super()._deserialize(value, attr, data, **kwargs)
        return deserialized_str.upper()


class MsgSchema(Schema):
    msg = fields.Str()


class CustomValuatorResultSchema(Schema):
    client_unit_id = fields.Str(required=True, allow_none=False)
    ph_final_gross_rent_annual_m2 = fields.Float(required=True, allow_none=False)
    ph_final_gross_rent_adj_factor = fields.Float(required=True, allow_none=False)


class CSVFile(fields.Field):
    CATEGORY_SEPARATOR = "."

    def __init__(
        self,
        schema: Schema,
        delimiter=",",
        quotechar='"',
        encoding="utf-8",
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.delimiter = delimiter
        self.quotechar = quotechar
        self.encoding = encoding
        self.schema = schema

    def _deserialize(
        self,
        value: FileStorage,
        attr: Optional[str],
        data: Optional[Mapping[str, Any]],
        **kwargs,
    ):
        """
        Example File:
            unit_id,custom_field,...
            GS20.00.01,1.0,...
            ...
        """
        try:
            decoded = StringIO(value.read().decode(self.encoding))
        except UnicodeDecodeError:
            raise ValidationError(
                f"The file is in the wrong encoding. The file must be encoded in {self.encoding}"
            )

        return [
            self.schema.load(r)
            for r in csv.DictReader(
                decoded, delimiter=self.delimiter, quotechar=self.quotechar
            )
        ]


class CustomValuatorResultFileDeserializer:
    HEADERS_ROW = 7
    FIELD_MAPPING = {
        "^Unit Identifier$": "client_unit_id",
        r"^Final Predicted Gross Rent \(.* yearly/m2\)$": "ph_final_gross_rent_annual_m2",
        "^Rent Calibrator Adjustment Factor$": "ph_final_gross_rent_adj_factor",
    }

    @classmethod
    def deserialize(cls, file: FileStorage) -> list[dict]:
        try:
            sheet = load_workbook(file)["Global Overview"]
        except BadZipFile:
            raise ValidationError("The file is not a valid Excel file")

        header_row = [cell.value for cell in sheet[cls.HEADERS_ROW]]
        field_indices = cls._get_field_indices(row=header_row)

        return [
            CustomValuatorResultSchema().load(
                {
                    cls.FIELD_MAPPING[field]: cls._clean_str(row[index])
                    for field, index in field_indices.items()
                }
            )
            for row in sheet.iter_rows(min_row=cls.HEADERS_ROW + 1, values_only=True)
            if any(row)  # sometimes there could be empty rows at the end
        ]

    @classmethod
    def _get_field_indices(cls, row: list[str]) -> dict[str, int]:
        header_row = [val.lower().strip(" \t") for val in row]

        result = {}
        for field_pattern in cls.FIELD_MAPPING.keys():
            for i, name in enumerate(header_row):
                if re.match(field_pattern.lower(), name):
                    result[field_pattern] = i
                    break
            else:
                raise ValidationError(
                    f'The file does not contain the expected header at row {cls.HEADERS_ROW}: "{field_pattern[1:-1]}"'
                )
        return result

    @staticmethod
    def _clean_str(value: Union[str, int]) -> Union[str, int]:
        if isinstance(value, str):
            return value.strip(" \t")
        return value


class GCSLinkArgs(Schema):
    language = EnumField(SUPPORTED_LANGUAGES, by_value=False, required=True)
    file_format = EnumField(SUPPORTED_OUTPUT_FILES, by_value=False, required=True)


class UnionField(fields.Field):
    """Field that deserializes multi-type input data to app-level objects."""

    def __init__(self, val_types: list[fields.Field]):
        self.valid_types = val_types
        super().__init__()

    def _deserialize(
        self, value: Any, attr: str = None, data: Mapping[str, Any] = None, **kwargs
    ):
        errors = []
        # iterate through the types being passed into UnionField via valid_types
        for field in self.valid_types:
            try:
                return field.deserialize(value, attr, data, **kwargs)
            except ValidationError as error:
                errors.append(error.messages)
        raise ValidationError(errors)
