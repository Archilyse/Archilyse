from collections import defaultdict

import click
from openpyxl import load_workbook

from common_utils.constants import UNIT_USAGE
from handlers.db import BuildingDBHandler, FloorDBHandler, UnitDBHandler
from handlers.ph_vector import PHResultVectorHandler

MAIN_SHEETS = ["Global Overview", "Rent Calibrator Details"]
COLUMN_TO_STYLE = {"Global Overview": 13, "Rent Calibrator Details": 9}
COLUMN_MAPPING = {4: "street", 5: "net_area", 6: "number_of_rooms", 7: "floor_number"}


@click.command()
@click.option("--site_id", prompt=True, required=True, type=int)
@click.option("--input_file", prompt=True, required=True, type=click.Path(exists=True))
def process_excel_file(site_id: int, input_file: str):
    data = get_relevant_data_from_db(site_id=site_id)

    wb = load_workbook(input_file)

    clean_sheets(wb)
    for sheet in MAIN_SHEETS:
        fill_worksheet(wb[sheet], data, relevant_column=COLUMN_TO_STYLE[sheet])

    wb.save("output_rent_ext.xlsx")


def fill_worksheet(sheet, data, relevant_column):
    new_rows = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        representative_unit = row[0]
        for client_id, unit_info in data[representative_unit].items():
            new_row = list(row)
            new_row[0] = client_id
            for i, column in COLUMN_MAPPING.items():
                new_row[i] = unit_info[column]
            new_rows.append(new_row)

    for row in new_rows:
        sheet.append(row)
        # Take row just inserted and the last column cell and change the style to Normal
        sheet.cell(row=sheet.max_row, column=relevant_column).style = "Normal"


def get_relevant_data_from_db(site_id: int) -> dict:
    # makes a query in raw sql to the db to select all units
    # that are relevant to the site_id
    units_info = UnitDBHandler.find(
        site_id=site_id,
        unit_usage=UNIT_USAGE.RESIDENTIAL.name,
        output_columns=["id", "client_id", "floor_id", "representative_unit_client_id"],
    )

    building_info = {
        b["id"]: b
        for b in BuildingDBHandler.find(
            site_id=site_id, output_columns=["id", "street"]
        )
    }

    floors_info = {
        f["id"]: {"building_id": f["building_id"], "floor_number": f["floor_number"]}
        for f in FloorDBHandler.find_in(
            building_id=building_info.keys(),
            output_columns=["id", "floor_number", "building_id"],
        )
    }
    basic_features = PHResultVectorHandler(site_id=site_id).basic_features
    info_by_representative_unit = defaultdict(dict)

    for unit_info in units_info:
        if unit_info["representative_unit_client_id"] == unit_info["client_id"]:
            # Ignore representative units, as we already have them on the spreadsheet
            continue
        client_id = unit_info["client_id"]
        building_id = floors_info[unit_info["floor_id"]]["building_id"]
        info_by_representative_unit[unit_info["representative_unit_client_id"]][
            client_id
        ] = {
            "floor_number": floors_info[unit_info["floor_id"]]["floor_number"],
            "street": building_info[building_id]["street"],
            "net_area": round(basic_features[client_id]["UnitBasics.net-area"]),
            "number_of_rooms": basic_features[client_id]["UnitBasics.number-of-rooms"],
        }
    return info_by_representative_unit


def clean_sheets(wb):
    # Remove sheets that are not in the list SHEETS_2_KEEP
    for sheet in wb.sheetnames:
        if sheet not in MAIN_SHEETS:
            wb.remove(wb[sheet])


if __name__ == "__main__":
    process_excel_file()
